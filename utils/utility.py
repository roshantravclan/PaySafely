import csv

from openpyxl import Workbook, load_workbook


class LocalStorage:
    def __init__(self, driver):
        self.driver = driver

    def set(self, key, value):
        self.driver.execute_script("window.localStorage.setItem(arguments[0], arguments[1]);", key, value)


def read_data_from_excel(file_name, sheet):
    data_list = []
    wb = load_workbook(file_name)
    sh = wb[sheet]

    row_ct = sh.max_row
    col_ct = sh.max_column

    for i in range(2, row_ct + 1):
        row = []
        for j in range(1, col_ct + 1):
            row.append(sh.cell(row=i, column=j).value)
        data_list.append(row)
    return data_list


def read_data_from_csv(file_name):
    data_list = []
    csv_data = open(file_name, "r")
    reader = csv.reader(csv_data)
    next(reader)
    for row in reader:
        data_list.append(row)
    return data_list
